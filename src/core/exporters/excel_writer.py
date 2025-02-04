from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import List

from models.store_aggregate import StoreAggregate
from models.item_aggregate import ItemAggregate

class ExcelExporter:
    output_path: Path

    def export(self, store_aggregates: List[StoreAggregate], item_aggregates: List[ItemAggregate]):
        wb = Workbook()

        store_sheet = wb.active
        store_sheet.title = "Butikker"
        self.__write_to_sheet__(store_sheet, store_aggregates, ["Butikknavn", "Totalt beløp", "Antall kvitteringer"])

        item_sheet = wb.create_sheet(title="Varer")
        self.__write_to_sheet__(item_sheet, item_aggregates, ["Varenavn", "Total antall", "Total beløp"])

        wb.save(self.output_path)
        print(f"Excel-fil lagret til: {self.output_path}")

    def __write_to_sheet__(self, sheet, data, headers):
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)

        for row_idx, item in enumerate(data, start=2):
            for col_idx, value in enumerate(item.__dict__.values(), start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

        for col_idx, _ in enumerate(headers, start=1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = 20
