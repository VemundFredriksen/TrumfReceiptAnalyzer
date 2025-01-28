from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import List

from models.butikk_aggregate import ButikkAggregate
from models.vare_aggregate import VareAggregate

@dataclass
class ExcelExporter:
    output_path: Path

    def lag_rapport(self, butikk_aggregater: List[ButikkAggregate], vare_aggregater: List[VareAggregate]):
        wb = Workbook()

        butikk_sheet = wb.active
        butikk_sheet.title = "Butikker"
        self._skriv_til_ark(butikk_sheet, butikk_aggregater, ["Butikknavn", "Totalt beløp", "Antall kvitteringer"])

        vare_sheet = wb.create_sheet(title="Varer")
        self._skriv_til_ark(vare_sheet, vare_aggregater, ["Varenavn", "Total antall", "Total beløp"])

        wb.save(self.output_path)
        print(f"Excel-fil lagret til: {self.output_path}")

    def _skriv_til_ark(self, sheet, data, headers):
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)

        for row_idx, item in enumerate(data, start=2):
            for col_idx, value in enumerate(item.__dict__.values(), start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

        for col_idx, _ in enumerate(headers, start=1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = 20
